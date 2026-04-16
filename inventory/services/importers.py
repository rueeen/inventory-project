from decimal import Decimal, InvalidOperation
import re

from openpyxl import load_workbook

from ..models import AcademicArea, Career, Equipment, StorageLocation, Subject, Supply


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_int(value, default=0):
    if value in (None, ""):
        return default

    try:
        text = str(value).strip().replace(",", ".")
        return int(float(text))
    except (TypeError, ValueError):
        return default


def parse_decimal(value):
    if value in (None, ""):
        return None

    try:
        text = str(value).strip().replace(",", ".")
        return Decimal(text)
    except (InvalidOperation, TypeError, ValueError):
        return None


def split_careers(text):
    if not text:
        return []

    raw = str(text).strip()

    if "," in raw or ";" in raw or "\n" in raw:
        parts = re.split(r"[,;\n]+", raw)
    else:
        parts = re.split(r"\s+y\s+|\s+e\s+", raw, flags=re.IGNORECASE)

    return [part.strip() for part in parts if part and part.strip()]


def split_subjects(text):
    if not text:
        return []

    parts = re.split(r"[,;\n]+", str(text))
    result = []

    for part in parts:
        item = part.strip()
        if not item:
            continue

        if " - " in item:
            code, name = item.split(" - ", 1)
            result.append((code.strip(), name.strip()))
        else:
            result.append((item, ""))

    unique = []
    seen = set()
    for code, name in result:
        if code not in seen:
            unique.append((code, name))
            seen.add(code)

    return unique


def resolve_condition(good_count, repairable_count, bad_count):
    good_count = parse_int(good_count, 0)
    repairable_count = parse_int(repairable_count, 0)
    bad_count = parse_int(bad_count, 0)

    if bad_count > 0:
        return "bad"
    if repairable_count > 0:
        return "repairable"
    return "good"


def get_portada_area_name(workbook):
    if "Portada" not in workbook.sheetnames:
        return ""

    ws = workbook["Portada"]
    return clean_text(ws["E5"].value)


def import_equipment_excel(file_obj, academic_area_name=None):
    wb = load_workbook(file_obj, data_only=True)

    if "Equipos" not in wb.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'Equipos'.")

    ws = wb["Equipos"]

    result = {
        "created": 0,
        "updated": 0,
        "errors": [],
    }

    area_name = academic_area_name or get_portada_area_name(wb)
    academic_area = None
    if area_name:
        academic_area, _ = AcademicArea.objects.get_or_create(name=area_name)

    for row_num, row in enumerate(ws.iter_rows(min_row=6, max_col=15, values_only=True), start=6):
        try:
            inventory_code = clean_text(row[0])
            name = clean_text(row[1])
            detailed_spec = clean_text(row[2])
            careers_text = row[3]
            subjects_text = row[4]
            location_name = clean_text(row[5])
            good_count = row[9]
            repairable_count = row[10]
            bad_count = row[11]
            unit_value_uf = row[12]
            observations = clean_text(row[14])

            if not inventory_code and not name:
                continue

            if not inventory_code:
                result["errors"].append(
                    f"Fila {row_num}: no se encontró un código de inventario válido."
                )
                continue

            if not name:
                result["errors"].append(
                    f"Fila {row_num}: no se encontró nombre de equipo."
                )
                continue

            location, _ = StorageLocation.objects.get_or_create(
                name=location_name or "Sin ubicación"
            )

            condition = resolve_condition(
                good_count, repairable_count, bad_count)

            career_objects = []
            for career_name in split_careers(careers_text):
                career, _ = Career.objects.get_or_create(name=career_name)
                career_objects.append(career)

            subject_objects = []
            for subject_code, subject_name in split_subjects(subjects_text):
                subject, _ = Subject.objects.get_or_create(
                    code=subject_code,
                    defaults={"name": subject_name},
                )

                if subject_name and subject.name != subject_name:
                    subject.name = subject_name
                    subject.save(update_fields=["name"])

                subject_objects.append(subject)

            defaults = {
                "name": name,
                "detailed_spec": detailed_spec,
                "storage_location": location,
                "condition": condition,
                "unit_value_uf": parse_decimal(unit_value_uf),
                "observations": observations,
            }

            if academic_area is not None:
                defaults["academic_area"] = academic_area

            equipment, created = Equipment.objects.update_or_create(
                inventory_code=inventory_code,
                defaults=defaults,
            )

            equipment.careers.set(career_objects)
            equipment.subjects.set(subject_objects)

            if created:
                result["created"] += 1
            else:
                result["updated"] += 1

        except Exception as e:
            result["errors"].append(f"Fila {row_num}: {str(e)}")

    return result


def import_supply_excel(file_obj, academic_area_name=None):
    wb = load_workbook(file_obj, data_only=True)

    if "Insumos" not in wb.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'Insumos'.")

    ws = wb["Insumos"]

    result = {
        "created": 0,
        "updated": 0,
        "errors": [],
    }

    area_name = academic_area_name or get_portada_area_name(wb)
    academic_area = None
    if area_name:
        academic_area, _ = AcademicArea.objects.get_or_create(name=area_name)

    for row_num, row in enumerate(ws.iter_rows(min_row=6, max_col=5, values_only=True), start=6):
        try:
            name = clean_text(row[0])
            detailed_spec = clean_text(row[1])
            location_name = clean_text(row[2])
            total_existing = parse_int(row[3], 0)
            observations = clean_text(row[4])

            if not name:
                continue

            location, _ = StorageLocation.objects.get_or_create(
                name=location_name or "Sin ubicación"
            )

            defaults = {
                "detailed_spec": detailed_spec,
                "storage_location": location,
                "total_existing": total_existing,
                "observations": observations,
            }

            if academic_area is not None:
                defaults["academic_area"] = academic_area

            _, created = Supply.objects.update_or_create(
                name=name,
                storage_location=location,
                defaults=defaults,
            )

            if created:
                result["created"] += 1
            else:
                result["updated"] += 1

        except Exception as e:
            result["errors"].append(f"Fila {row_num}: {str(e)}")

    return result
